from django.shortcuts import render, redirect, get_object_or_404
from .models import Printer, PrinterLog, Employee, Equipment
from .forms import EmployeeForm, EquipmentForm, CSVImportForm, PrinterForm
from django.contrib import messages
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import csv
from .models import Employee  # Import modelu Employee z Twojej aplikacji
from .models import Employee
from .forms import EmployeeForm


# View to add a new printer
def add_printer(request):
    if request.method == 'POST':
        form = PrinterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('printer_list')  # Redirect to a list of printers after saving
    else:
        form = PrinterForm()
    return render(request, 'core/add_printer.html', {'form': form})

def printer_list(request):
    printers = Printer.objects.all()
    return render(request, 'core/printer_list.html', {'printers': printers})

def edit_printer(request, pk):
    printer = get_object_or_404(Printer, pk=pk)
    if request.method == 'POST':
        form = PrinterForm(request.POST, instance=printer)
        if form.is_valid():
            form.save()
            return redirect('printer_list')
    else:
        form = PrinterForm(instance=printer)
    return render(request, 'core/edit_printer.html', {'form': form})

def delete_printer(request, pk):
    printer = get_object_or_404(Printer, pk=pk)
    if request.method == 'POST':
        printer.delete()
        return redirect('printer_list')
    return render(request, 'core/delete_printer.html', {'printer': printer})

def printer_detail(request, pk):
    from itertools import chain

    printer = get_object_or_404(Printer, pk=pk)
    snmp_data = {}

    # Mapowanie zmiennych do ich głównych i alternatywnych OID
    oid_map = {
        "Model": [".1.3.6.1.4.1.1602.1.1.1.1.0"],
        "Serial Number": [".1.3.6.1.4.1.1602.1.2.1.8.1.3.1.1"],
        "Wydruki Czarno Białe": [
            ".1.3.6.1.4.1.1602.1.11.1.3.1.4.109",
            ".1.3.6.1.4.1.1602.1.11.2.1.1.3.5",
            ".1.3.6.1.2.1.43.10.2.1.4.1.1",
        ],
        "Wydruki kolorowe": [
            ".1.3.6.1.4.1.1602.1.11.1.4.1.4.123",
            ".1.3.6.1.4.1.1602.1.11.2.1.1.3.3",
        ],
        "Suma wydruków": [
            ".1.3.6.1.4.1.1602.1.11.1.4.1.4.101",
            ".1.3.6.1.4.1.1602.1.11.2.1.1.3.1",
        ],
        "Suma skanów": [".1.3.6.1.4.1.1602.1.11.1.4.1.4.501"],
    }

    if request.method == "POST":
        for key, oids in oid_map.items():
            success = False
            for oid in oids:
                try:
                    # Uruchomienie komendy SNMP dla danego OID
                    snmp_output = subprocess.check_output(
                        [
                            "C:/Tools/SnmpGet.exe",
                            "-c:public",
                            f"-r:{printer.ip_address}",
                            f"-o:{oid}",
                        ],
                        universal_newlines=True,
                    ).strip()

                    # Wyciągnięcie wartości z wyniku
                    for line in snmp_output.splitlines():
                        if "Value=" in line:
                            snmp_data[key] = line.split("Value=")[-1].strip()
                            success = True
                            break

                    # Jeśli dane zostały pobrane, nie próbuj alternatywnych OID
                    if success:
                        break
                except subprocess.CalledProcessError as e:
                    snmp_data[key] = f"Błąd SNMP: {e.returncode}"
                except FileNotFoundError:
                    snmp_data[key] = "Nie znaleziono SnmpGet.exe. Sprawdź konfigurację PATH."
                except Exception as e:
                    snmp_data[key] = f"Błąd: {str(e)}"

            # Jeśli żaden OID nie zwrócił wartości
            if not success:
                snmp_data[key] = "Brak danych lub nieobsługiwany model drukarki"

    return render(request, "core/printer_detail.html", {"printer": printer, "snmp_data": snmp_data})

def employee_list(request):
    employees = Employee.objects.all()
    return render(request, 'core/employee_list.html', {'employees': employees})

def add_employee(request):
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'core/add_employee.html', {'form': form})

def add_equipment(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('equipment_list')
    else:
        form = EquipmentForm()
    return render(request, 'core/add_equipment.html', {'form': form})

def equipment_list(request):
    equipment = Equipment.objects.all()
    return render(request, 'core/equipment_list.html', {'equipment': equipment})

def edit_equipment(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            return redirect('equipment_list')
    else:
        form = EquipmentForm(instance=equipment)
    return render(request, 'core/edit_equipment.html', {'form': form})

def delete_equipment(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        equipment.delete()
        return redirect('equipment_list')
    return render(request, 'core/delete_equipment.html', {'equipment': equipment})

def edit_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'core/edit_employee.html', {'form': form})

def delete_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('employee_list')
    return render(request, 'core/delete_employee.html', {'employee': employee})

def index(request):
    return render(request, 'core/index.html')

def import_csv(request):
    if request.method == 'POST':
        csv_file = request.FILES['csv_file']
        table_name = request.POST.get('table_name')  # Pobranie wartości table_name z formularza
        delimiter = request.POST.get('delimiter', ',')  # Pobranie separatora z formularza (domyślnie ',')

        # Sprawdzenie formatu pliku
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'File is not in CSV format.')
            return redirect('import_csv')

        data = csv_file.read().decode('utf-8-sig')  # Usunięcie BOM
        csv_reader = csv.DictReader(data.splitlines(), delimiter=delimiter)

        # Obsługa importu do różnych tabel
        if table_name == 'equipment':
            for row in csv_reader:
                # Sprawdzenie wymaganych kolumn
                if 'assigned_to' not in row or 'type' not in row or 'model' not in row:
                    messages.error(request, "Missing required columns in CSV.")
                    return redirect('import_csv')

                # Przetwarzanie pola assigned_to
                assigned_to_employeename = row.get('assigned_to', "Unassigned").strip()
                if assigned_to_employeename.lower() == "unassigned":
                    assigned_to_employee = None
                else:
                    try:
                        # Rozdzielenie imienia i nazwiska oraz pobranie użytkownika
                        first_name, last_name = assigned_to_employeename.split()
                        assigned_to_employee = Employee.objects.get(first_name=first_name, last_name=last_name)
                    except (Employee.DoesNotExist, ValueError):
                        messages.error(
                            request,
                            f"Employee '{assigned_to_employeename}' does not exist or has an invalid format. Skipping entry."
                        )
                        continue  # Pomijamy ten wiersz

                # Tworzenie obiektu Equipment
                try:
                    Equipment.objects.create(
                        type=row.get('type', 'Unknown'),
                        model=row.get('model', 'Unknown'),
                        serial_number=row.get('serial_number', 'Unknown'),
                        condition=row.get('condition', 'Unknown'),
                        assigned_to=assigned_to_employee,
                    )
                except Exception as e:
                    messages.error(request, f"Error creating equipment entry: {e}")
                    continue  # Pomijamy wiersz z błędem
            messages.success(request, 'Equipment data imported successfully!')

        elif table_name == 'employee':
            for row in csv_reader:
                try:
                    # Tworzenie obiektu Employee
                    Employee.objects.create(
                        first_name=row['first_name'],
                        last_name=row['last_name'],
                        email=row['email'],
                    )
                except Exception as e:
                    messages.error(request, f"Error creating employee: {e}")
                    continue  # Pomijamy wiersz z błędem
            messages.success(request, 'Employee data imported successfully!')

        elif table_name == 'printer':
            for row in csv_reader:
                try:
                    # Tworzenie obiektu Printer
                    Printer.objects.create(
                        ip_address=row['ip_address'],
                        hostname=row['hostname'],
                        model=row['model'],
                        serial_number=row['serial_number'],
                        mac_address=row['mac_address'],
                        status=row['status'],
                    )
                except Exception as e:
                    messages.error(request, f"Error creating printer: {e}")
                    continue  # Pomijamy wiersz z błędem
            messages.success(request, 'Printer data imported successfully!')

        else:
            messages.error(request, 'Invalid table selected.')

        return redirect('import_csv')

    return render(request, 'core/import_csv.html')


def download_template(request, model_name):
    model_fields = {
        'equipment': ['type', 'model', 'serial_number', 'condition', 'assigned_to'],
        'employee': ['first_name', 'last_name', 'email', 'role'],
        'printer': ['ip_address', 'hostname', 'model', 'serial_number', 'mac_address', 'status'],
    }

    if model_name not in model_fields:
        return HttpResponse("Invalid model name", status=400)

    # Tworzenie pliku Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{model_name.capitalize()} Template"

    # Dodanie nagłówków
    fields = model_fields[model_name]
    for col_num, field in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = field

    # Generowanie odpowiedzi z plikiem
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{model_name}_template.xlsx"'
    wb.save(response)
    return response


def user_list(request):
    users = Employee.objects.all()  # lub inny model, jeśli został zmieniony
    return render(request, 'core/user_list.html', {'users': users})
